"""
Processador de planilhas Excel.
Responsável pela leitura, processamento e escrita de dados em arquivos .xlsm.
"""

import openpyxl
from scrapers.droga_raia import consultar_droga_raia
from scrapers.drogarias_pacheco import consultar_pacheco
from scrapers.super_nosso import consultar_super_nosso


def processar_excel_com_scraping(driver, caminho_entrada, caminho_saida, flags):
    """
    Processa arquivo Excel com scraping de preços em múltiplas farmácias.
    
    Args:
        driver: WebDriver Selenium
        caminho_entrada: Caminho do arquivo Excel de entrada
        caminho_saida: Caminho onde salvar o resultado
        flags: Dict com flags de ativação {'raia': bool, 'pacheco': bool, 'supernosso': bool}
    
    Raises:
        Exception: Qualquer erro durante o processamento
    """
    wb = openpyxl.load_workbook(caminho_entrada, keep_vba=True)
    ws = wb.active

    for row in range(5, ws.max_row + 1):
        ean_val = ws[f'A{row}'].value
        if not ean_val:
            continue

        ean = str(ean_val).split('.')[0].strip()
        print(f"🔎 Buscando EAN: {ean} na linha {row}")

        if flags.get('raia'):
            valor, vendedor = consultar_droga_raia(driver, ean)
            if valor:
                if vendedor == "PROPRIO":
                    ws[f'D{row}'] = valor
                    ws[f'D{row}'].number_format = '#,##0.00'
                else:
                    ws[f'R{row}'] = valor
                    ws[f'R{row}'].number_format = '#,##0.00'
                    ws[f'S{row}'] = vendedor

        if flags.get('pacheco'):
            valor, seller = consultar_pacheco(driver, ean)
            if valor:
                if seller == "PROPRIO":
                    ws[f'E{row}'] = valor
                    ws[f'E{row}'].number_format = '#,##0.00'
                else:
                    if not ws[f'R{row}'].value:
                        ws[f'R{row}'] = valor
                        ws[f'R{row}'].number_format = '#,##0.00'
                        ws[f'S{row}'] = f"{seller} (Pacheco)"

        if flags.get('supernosso'):
            valor, seller = consultar_super_nosso(driver, ean)
            if valor:
                ws[f'F{row}'] = valor
                ws[f'F{row}'].number_format = '#,##0.00'

        wb.save(caminho_saida)
