from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import os
import time
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'processamento'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

driver = None

def iniciar_driver():
    chrome_options = Options()
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=chrome_options)


def fechar_cookies_raia():
    seletores = ["onetrust-accept-btn-handler", "//button[contains(text(), 'Aceitar')]"]
    for s in seletores:
        try:
            botao = driver.find_element(By.XPATH, s) if s.startswith("//") else driver.find_element(By.ID, s)
            driver.execute_script("arguments[0].click();", botao)
            time.sleep(1)
            break
        except Exception:
            continue

def tratar_valor_raia(texto):
    try:
        limpo = texto.lower().split('cada')[0].split('un')[0].strip()
        return float(limpo.replace('r$', '').replace('\xa0', '').replace('.', '').replace(',', '.'))
    except Exception:
        return None

def consultar_droga_raia(ean):
    try:
        driver.get(f"https://www.drogaraia.com.br/search?w={ean}")
        time.sleep(random.uniform(9, 11)) 
        fechar_cookies_raia()

        if "não encontramos resultados" in driver.page_source.lower():
            return None, None

        seletores_imagem = [
            'img[data-testid="product-image"]',
            'a[data-qa="product-card-link"] img',
            'a[data-qa="product-card-link"]'
        ]
        
        for sel in seletores_imagem:
            try:
                elemento = driver.find_element(By.CSS_SELECTOR, sel)
                driver.execute_script("arguments[0].click();", elemento)
                time.sleep(10)
                break
            except Exception:
                continue
        else:
            return None, None 

        try:
            preco_elem = driver.find_element(By.CSS_SELECTOR, 'span[class*="price-pdp-content"], span.eVpKuL')
            preco_final = tratar_valor_raia(preco_elem.text)
            
            texto_pagina = driver.find_element(By.TAG_NAME, 'body').text.lower()
            termos_raia = ["vendido e entregue por raia", "vendido por droga raia", "entregue por droga raia", "por raia"]
            
            if any(termo in texto_pagina for termo in termos_raia):
                return preco_final, "PROPRIO"
            
            vendedor = driver.find_element(By.CSS_SELECTOR, 'button[data-testid="seller-button"]').text.strip()
            return preco_final, vendedor
            
        except Exception:
            return preco_final if 'preco_final' in locals() else None, "Marketplace Raia" if 'preco_final' in locals() else None

    except Exception:
        return None, None


def exterminar_modais_pacheco():
    js_ninja = """
    var btnCookie = document.querySelector('div.dp-bar-dismiss');
    if(btnCookie) { btnCookie.click(); }
    var btnCookieAlt = document.getElementById('onetrust-accept-btn-handler');
    if(btnCookieAlt) { btnCookieAlt.click(); }
    var btnPbm = document.querySelector('button.btn-close[data-bs-dismiss="modal"]');
    if(btnPbm) { btnPbm.click(); }
    var btnVtex = document.querySelector('.vtex-modal-layout-0-x-closeButton');
    if(btnVtex) { btnVtex.click(); }
    """
    try:
        driver.execute_script(js_ninja)
    except Exception:
        pass

def tratar_valor_pacheco(texto):
    try:
        if not texto: return None
        limpo = texto.lower().split('cada')[0].split('un')[0].split('por')[-1].strip()
        return float(limpo.replace('r$', '').replace('\xa0', '').replace('.', '').replace(',', '.'))
    except Exception:
        return None

def consultar_pacheco(ean):
    try:
        driver.get(f"https://www.drogariaspacheco.com.br/pesquisa?q={ean}")
        time.sleep(random.uniform(6, 8))
        exterminar_modais_pacheco()

        texto_pesquisa = driver.find_element(By.TAG_NAME, 'body').text.lower()
        if "não encontramos" in texto_pesquisa or "ops!" in texto_pesquisa:
            return None, None

        js_pegar_link = """
        var links = document.querySelectorAll('a[href$="/p"]');
        for(var i=0; i < links.length; i++) {
            var rect = links[i].getBoundingClientRect();
            if(rect.top > 0 && rect.top < window.innerHeight && rect.width > 50 && rect.height > 50) {
                if(window.getComputedStyle(links[i]).visibility !== 'hidden') {
                    return links[i].href;
                }
            }
        }
        var galeria = document.querySelector('div[id^="gallery"], div[class*="search-result"]');
        if(galeria) {
            var linkGaleria = galeria.querySelector('a[href$="/p"]');
            if(linkGaleria) return linkGaleria.href;
        }
        return null;
        """
        link_produto = driver.execute_script(js_pegar_link)

        if link_produto:
            driver.get(link_produto) 
            time.sleep(random.uniform(5, 7)) 
            exterminar_modais_pacheco()
        else:
            return None, None

        seletores_preco = [
            '[class*="sellingPriceValue"]',
            '.vtex-store-components-3-x-sellingPriceValue',
            '.vtex-product-price-1-x-sellingPriceValue',
            '[class*="sellingPrice"]',
            '.valor-por',
            'strong.skuBestPrice'
        ]

        preco_final = None
        for sel_p in seletores_preco:
            try:
                preco_elem = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, sel_p))
                )
                texto_preco = preco_elem.text.strip()
                if texto_preco:  
                    preco_final = tratar_valor_pacheco(texto_preco)
                    if preco_final: 
                        break
            except Exception:
                continue

        if not preco_final:
            return None, None
            
        try:
            try:
                vendedor_mkt = driver.find_element(By.CSS_SELECTOR, 'a[data-bs-target="#rnk-comp-modal-seller"]')
                nome_vendedor = vendedor_mkt.text.strip()
                if nome_vendedor and "pacheco" not in nome_vendedor.lower():
                    return preco_final, nome_vendedor
            except Exception:
                pass 
            
            try:
                vendedor_elem = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.vtex-store-components-3-x-sellerName, [class*="sellerName"]'))
                )
                nome_vendedor = vendedor_elem.text.strip()
                if "pacheco" in nome_vendedor.lower():
                    return preco_final, "PROPRIO"
                elif nome_vendedor != "":
                    return preco_final, nome_vendedor
            except Exception:
                pass
                
            texto_pagina = driver.find_element(By.TAG_NAME, 'body').text.lower()
            if "vendido e entregue por" in texto_pagina:
                trecho = texto_pagina.split("vendido e entregue por")[1].split('\n')[0].strip()
                if "pacheco" in trecho:
                    return preco_final, "PROPRIO"
                else:
                    return preco_final, trecho.title()
                    
            if "entregue por drogaria pacheco" in texto_pagina or "vendido por drogaria pacheco" in texto_pagina:
                 return preco_final, "PROPRIO"
                 
            return preco_final, "PROPRIO"
        except Exception:
            return preco_final, "PROPRIO"

    except Exception:
        return None, None


def tratar_valor_super_nosso(texto):
    try:
        if not texto: return None
        limpo = texto.lower().replace('r$', '').replace('\n', '').replace(' ', '').replace('\xa0', '').strip()
        limpo = limpo.replace('.', '')
        return float(limpo.replace(',', '.'))
    except Exception:
        return None

def consultar_super_nosso(ean):
    try:
        driver.get(f"https://www.supernosso.com/{ean}")
        time.sleep(random.uniform(6, 8))

        texto_pesquisa = driver.find_element(By.TAG_NAME, 'body').text.lower()
        if "não encontramos" in texto_pesquisa or "nenhum resultado" in texto_pesquisa:
            return None, None

        js_pegar_link_sn = """
        var links = document.querySelectorAll('a[href$="/p"]');
        for(var i=0; i < links.length; i++) {
            if(links[i].href && links[i].href.includes('supernosso.com')) {
                return links[i].href;
            }
        }
        var article = document.querySelector('article.vtex-product-summary-2-x-element');
        if(article) {
            var link = article.querySelector('a');
            if(link && link.href) return link.href;
        }
        return null;
        """
        link_produto = driver.execute_script(js_pegar_link_sn)

        if link_produto:
            print(f"🔗 [Super Nosso] Entrando no produto: {link_produto}")
            driver.get(link_produto)
            time.sleep(random.uniform(5, 7))
        else:
            print("❌ [Super Nosso] Produto não encontrado na busca.")
            return None, None

        js_pegar_preco = """
        // Procura a caixa principal do preço de venda
        var caixaPreco = document.querySelector('[class*="sellingPrice"]');
        if(caixaPreco) {
            // Pega o número inteiro (ex: 27) e os centavos (ex: 99)
            var intElem = caixaPreco.querySelector('[class*="currencyInteger"]');
            var fracElem = caixaPreco.querySelector('[class*="currencyFraction"]');
            
            if(intElem && fracElem) {
                // Junta os dois com uma vírgula
                return intElem.innerText.trim() + ',' + fracElem.innerText.trim();
            }
            // Se não achar os pedaços separados, devolve o texto inteiro da caixa
            return caixaPreco.innerText;
        }
        return null;
        """
        texto_preco = driver.execute_script(js_pegar_preco)
        
        if not texto_preco:
            try:
                elem = driver.find_element(By.CSS_SELECTOR, '[class*="sellingPriceValue"]')
                texto_preco = elem.text
            except:
                pass

        if texto_preco:
            preco_final = tratar_valor_super_nosso(texto_preco)
            if preco_final:
                print(f"💰 [Super Nosso] Preço capturado: R$ {preco_final}")
                return preco_final, "PROPRIO"
                
        return None, None

    except Exception as e:
        return None, None


@app.route('/pesquisar', methods=['POST'])
def pesquisar():
    global driver
    dados = request.json
    ean = dados.get('ean')
    lojas = dados.get('lojas', [])
    
    if not driver: 
        driver = iniciar_driver()
        
    resultados = {'raia': '-', 'pacheco': '-', 'supernosso': '-'}
    
    if 'raia' in lojas:
        valor, vend = consultar_droga_raia(ean)
        resultados['raia'] = f"R$ {valor:.2f}" if valor else "Não enc."
        
    if 'pacheco' in lojas:
        valor, vend = consultar_pacheco(ean)
        resultados['pacheco'] = f"R$ {valor:.2f}" if valor else "Não enc."
        
    if 'supernosso' in lojas:
        valor, vend = consultar_super_nosso(ean)
        resultados['supernosso'] = f"R$ {valor:.2f}" if valor else "Não enc."
        
    return jsonify(resultados)


@app.route('/processar-excel', methods=['POST'])
def processar_excel():
    global driver
    file = request.files['file']
    
    fleg_raia = str(request.form.get('raia')).lower() in ['true', 'on', '1', 'yes']
    fleg_pacheco = str(request.form.get('pacheco')).lower() in ['true', 'on', '1', 'yes']
    fleg_supernosso = str(request.form.get('supernosso', request.form.get('checkSuperNosso'))).lower() in ['true', 'on', '1', 'yes']
    
    print(f"\n🚀 Flags de Ativação -> Raia: {fleg_raia} | Pacheco: {fleg_pacheco} | Super Nosso: {fleg_supernosso}")
    
    caminho_entrada = os.path.join(UPLOAD_FOLDER, "entrada.xlsm")
    caminho_saida = os.path.join(UPLOAD_FOLDER, "RESULTADO_FINAL.xlsm")
    file.save(caminho_entrada)

    if not driver: 
        driver = iniciar_driver()

    wb = openpyxl.load_workbook(caminho_entrada, keep_vba=True)
    ws = wb.active

    for row in range(5, ws.max_row + 1):
        ean_val = ws[f'A{row}'].value
        if not ean_val: continue
        
        ean = str(ean_val).split('.')[0].strip()
        print(f"🔎 Buscando EAN: {ean} na linha {row}")

        if fleg_raia:
            valor, vendedor = consultar_droga_raia(ean)
            if valor:
                if vendedor == "PROPRIO":
                    ws[f'D{row}'] = valor
                    ws[f'D{row}'].number_format = '#,##0.00'
                else:
                    ws[f'R{row}'] = valor
                    ws[f'R{row}'].number_format = '#,##0.00'
                    ws[f'S{row}'] = vendedor

        if fleg_pacheco:
            valor, seller = consultar_pacheco(ean)
            if valor:
                if seller == "PROPRIO":
                    ws[f'E{row}'] = valor
                    ws[f'E{row}'].number_format = '#,##0.00'
                else:
                    if not ws[f'R{row}'].value:
                        ws[f'R{row}'] = valor
                        ws[f'R{row}'].number_format = '#,##0.00'
                        ws[f'S{row}'] = f"{seller} (Pacheco)"

        if fleg_supernosso:
            valor, seller = consultar_super_nosso(ean)
            if valor:
                ws[f'F{row}'] = valor
                ws[f'F{row}'].number_format = '#,##0.00'

        wb.save(caminho_saida)

    return jsonify({"status": "concluido"})


@app.route('/download')
def download():
    return send_file(os.path.join(UPLOAD_FOLDER, "RESULTADO_FINAL.xlsm"), as_attachment=True)

if __name__ == '__main__':
    app.run(port=5000)